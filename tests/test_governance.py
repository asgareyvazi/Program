# ============================================================================
# GOVERNANCE TEST SUITE
# File: tests/test_governance.py
# Audit items (P2 / buyer Q4):
#   - Backup/restore round-trip (plain + encrypted at rest)
#   - Project revision snapshots (restore a previous version)
#   - Secrets manager round-trip
#   - Engineering calculation register correctness
#
# Run:  python3 tests/test_governance.py
# Exit code 0 = all pass.
# ============================================================================

import json
import os
import sqlite3
import sys
import tempfile

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

_PASS = 0
_FAIL = 0


def ok(cond, label, extra=""):
    global _PASS, _FAIL
    if cond:
        _PASS += 1
        print(f"  ✔ {label}")
    else:
        _FAIL += 1
        print(f"  ✘ {label} {extra}")


def test_encrypted_backup():
    print("\n[1] ENCRYPTED BACKUP / RESTORE (encryption at rest)")
    from pathlib import Path
    import backup_restore as br
    tmp = tempfile.mkdtemp(prefix="drl_backup_test_")
    app_dir = Path(tmp) / "app"
    bak_dir = Path(tmp) / "backups"
    os.makedirs(app_dir, exist_ok=True)
    os.makedirs(bak_dir, exist_ok=True)
    br.APP_DIR = app_dir
    br.BACKUP_DIR = bak_dir
    # seed a fake DB
    con = sqlite3.connect(os.path.join(app_dir, "procedures.db"))
    con.execute("CREATE TABLE t (x)")
    con.execute("INSERT INTO t VALUES (42)")
    con.commit()
    con.close()

    # plain backup
    b1 = br.create_backup(tag="plain")
    ok(b1 and b1.is_dir(), "plain backup folder created")
    # encrypted backup
    b2 = br.create_backup(tag="enc", password="s3cret!")
    ok(b2 and b2.suffix == ".enc", "encrypted .enc archive created")
    ok(b2 and not b2.with_name(b2.name[:-4]).exists(),
       "plain folder removed after encryption")
    # list shows encrypted flag
    lst = br.list_backups()
    enc = [b for b in lst if b.get("encrypted")]
    ok(len(enc) == 1, "list_backups marks encrypted entry")
    # wrong password fails
    res = br.restore_backup(b2.name, password="wrong")
    ok(res.get("error") and "password" in res["error"],
       "wrong password rejected")
    # correct password restores
    con = sqlite3.connect(os.path.join(app_dir, "procedures.db"))
    con.execute("INSERT INTO t VALUES (99)")
    con.commit()
    con.close()
    res = br.restore_backup(b2.name, password="s3cret!")
    ok("procedures.db" in res and res["procedures.db"],
       "encrypted restore succeeded")
    con = sqlite3.connect(os.path.join(app_dir, "procedures.db"))
    vals = [r[0] for r in con.execute("SELECT x FROM t")]
    con.close()
    ok(vals == [42], f"restored DB content (got {vals})")

    # cleanup
    import shutil
    shutil.rmtree(tmp, ignore_errors=True)
    import backup_restore as br2
    br2.APP_DIR = Path.home() / ".drilling_program"
    br2.BACKUP_DIR = br2.APP_DIR / "backups"


def test_project_revisions():
    print("\n[2] PROJECT REVISION SNAPSHOTS (buyer Q4)")
    from drilling_database import DrillingProjectDatabase
    tmp = tempfile.mkdtemp(prefix="drl_proj_test_")
    dbp = os.path.join(tmp, "projects.db")
    db = DrillingProjectDatabase(db_path=dbp)

    from dataclasses import dataclass, field

    @dataclass
    class CI:
        well_name: str = ""
        field_name: str = ""
        operator_name: str = ""

    @dataclass
    class WI:
        well_type: str = ""

    @dataclass
    class P:
        company_info: CI = field(default_factory=CI)
        well_info: WI = field(default_factory=WI)

    p1 = P(CI(well_name="Well A", field_name="F", operator_name="Op"))
    pid = db.save_project(p1, "Project A")
    # second save = revision 2
    p2 = P(CI(well_name="Well B", field_name="F", operator_name="Op"))
    db.save_project(p2, "Project A")
    revs = db.list_revisions(pid)
    ok(len(revs) == 2, f"two revisions stored (got {len(revs)})")
    # restore revision 1
    ok(db.restore_revision(pid, 1), "restore revision 1")
    data = db.get_revision_data(pid, 1)
    ok(data and data["company_info"]["well_name"] == "Well A",
       "revision 1 content restored (Well A)")
    cur = db.load_project(pid)
    ok(cur and cur["data"]["company_info"]["well_name"] == "Well A",
       "project_data replaced by revision 1")
    hist = db.get_history(pid)
    ok(any(h["action"] == "Restored" for h in hist),
       "restore recorded in history")
    db.close()
    import shutil
    shutil.rmtree(tmp, ignore_errors=True)


def test_secrets():
    print("\n[3] SECRETS MANAGER round-trip")
    from pathlib import Path
    from backup_restore import SecretsManager
    tmp = tempfile.mkdtemp(prefix="drl_secrets_")
    mgr = SecretsManager(service="TestSvc")
    mgr._fallback_file = Path(tmp) / "secrets.json"
    key = mgr.set_secret("api_key", "sk-test-123")
    ok(key is not None and len(key) > 0, "secret stored")
    ok(mgr.get_secret("api_key") == "sk-test-123", "secret retrieved")
    ok(mgr.delete_secret("api_key"), "secret deleted")
    ok(mgr.get_secret("api_key") in ("", None), "secret gone after delete")
    import shutil
    shutil.rmtree(tmp, ignore_errors=True)


def test_register():
    print("\n[4] ENGINEERING CALCULATION REGISTER")
    from engineering_register import compute_register, register_markdown
    from engineering_deep import deep_verify_markdown
    values = {
        "mud_weight": "12", "td_depth": "10000", "hole_size": "12.25",
        "pipe_od": "5", "casing_size": "9.625", "casing_wall": "0.472",
        "casing_yield": "110000", "casing_depth": "8000",
        "fracture_gradient": "16.0", "formation_pressure": "11.0",
        "sidpp": "400", "bop_wp": "10000", "flow_rate": "500",
        "yield_point": "20", "plastic_viscosity": "25", "trip_speed": "60",
        "pit_gain": "20", "annular_capacity": "0.045",
        "total_cost": "12000000", "total_days": "45",
        "n_index": "0.6", "k_index": "120", "yield_stress": "8",
        "burst_load": "9000", "collapse_load": "6000", "axial_load": "400000",
        "wob": "25", "rpm": "100",
    }
    rows = compute_register(values)
    ok(len(rows) >= 15, f"register rows >= 15 (got {len(rows)})")
    md = register_markdown(rows, "the Operator")
    ok("ENGINEERING CALCULATION REGISTER" in md, "register heading")
    ok("0.052" in md, "formula shown")
    ok("API 5C3" in md, "standard source shown")
    # deep verification section
    dmd = deep_verify_markdown(values, None, "the Operator")
    ok("DEEP ENGINEERING VERIFICATION" in dmd, "deep verify heading")
    ok("Herschel-Bulkley" in dmd, "HB hydraulics shown")
    ok("von Mises" in dmd, "triaxial shown")
    # ROP calibration flow
    from engineering_deep import ROPCalibrator
    rc = ROPCalibrator()
    offset = [
        {"wob": 20, "rpm": 90, "depth": 5000, "mw": 11, "rop_actual": 35},
        {"wob": 25, "rpm": 100, "depth": 8000, "mw": 11.5, "rop_actual": 28},
        {"wob": 30, "rpm": 110, "depth": 11000, "mw": 12, "rop_actual": 20},
    ]
    rc.calibrate(offset)
    dmd2 = deep_verify_markdown(values,
                                {"k": rc.k, "a": 1.0, "b": 0.6, "c": 0.00005,
                                 "d": -0.05, "n_points": len(offset)},
                                "the Operator")
    ok("Calibrated from **3** offset data" in dmd2, "calibration shown")
    ok("Predicted ROP" in dmd2 or "ROP (ft/hr)" in dmd2, "ROP table shown")


def test_structured_steps():
    print("\n[6] STRUCTURED STEP MODEL (hold/witness/precondition)")
    import tempfile
    from procedures_db import ProcedureDatabase
    tmp = tempfile.mkdtemp(prefix="drl_steps_")
    dbp = os.path.join(tmp, "procedures.db")
    db = ProcedureDatabase(db_path=dbp)
    cat = db.add_category("Test")
    pid = db.add_procedure("Test Procedure", cat)
    # add a structured step with hold + witness points
    db.add_step(pid, "Run casing to shoe depth.",
                precondition="Hole in good condition; returns established",
                acceptance="Casing landed; no drag over 50 klbf",
                hold_point=True, witness_point=True)
    db.add_step(pid, "Circulate and condition mud - verify returns.")
    steps = db.get_steps(pid)
    ok(len(steps) == 2, "two steps stored")
    s0 = steps[0]
    ok(s0.hold_point and s0.witness_point, "hold+witness flags round-trip")
    ok("good condition" in s0.precondition, "precondition round-trip")
    ok("no drag" in s0.acceptance, "acceptance round-trip")
    # update step
    db.update_step(steps[1].id, acceptance="Returns clean; no losses")
    s1 = db.get_steps(pid)[1]
    ok("Returns clean" in s1.acceptance, "acceptance update round-trip")
    # auto-structure parser
    from structured_steps import structure_step
    st = structure_step(
        "3.1 Pick up off bottom and verify string is free - hold point "
        "required before continuing, witness by company rep")
    ok(bool(st.action), "action extracted")
    ok(st.hold_point, "hold point auto-detected")
    ok("verify" in st.acceptance.lower(), "acceptance auto-detected")
    ok(st.witness_point, "witness point auto-detected")
    import shutil
    shutil.rmtree(tmp, ignore_errors=True)


def test_procedure_linking():
    print("\n[7] PROCEDURE <-> WELL/RISK LINKING + ROLE")
    import tempfile
    from procedures_db import ProcedureDatabase
    tmp = tempfile.mkdtemp(prefix="drl_link_")
    dbp = os.path.join(tmp, "procedures.db")
    db = ProcedureDatabase(db_path=dbp)
    cat = db.add_category("Test")
    pid = db.add_procedure("Casing Run Proc", cat,
                           linked_well_id="W-100",
                           linked_section="12.25 in hole",
                           linked_risk_ids="[3, 7]")
    # role round-trip on steps
    db.add_step(pid, "Supervise casing running", role="Supervisor",
                hold_point=True)
    s = db.get_steps(pid)[0]
    ok(s.role == "Supervisor", "step role round-trip")
    # links round-trip
    rec = db.get_procedure(pid)
    ok(rec.linked_well_id == "W-100", "linked well round-trip")
    ok(rec.linked_section == "12.25 in hole", "linked section round-trip")
    ok("3" in rec.linked_risk_ids and "7" in rec.linked_risk_ids,
       "linked risks round-trip")
    ok(db.get_links(pid)["well_id"] == "W-100", "get_links well")
    ok(db.get_links(pid)["risk_ids"] == [3, 7], "get_links risks")
    # query procedures_for_well
    found = db.procedures_for_well("W-100")
    ok(len(found) == 1 and found[0].name == "Casing Run Proc",
       "procedures_for_well finds linked procedure")
    ok(db.procedures_for_well("OTHER") == [], "other well -> none")
    # update links
    db.link_well(pid, "W-200", "8.5 in section")
    db.link_risks(pid, [1])
    rec2 = db.get_procedure(pid)
    ok(rec2.linked_well_id == "W-200", "link_well update")
    ok(rec2.linked_risk_ids == "[1]", "link_risks update")
    # regression: update_procedure with category_id + links together
    # (bindings bug fix — params must match placeholders)
    cat2 = db.add_category("Second Cat")
    db.update_procedure(pid, name="Renamed Proc", category_id=cat2,
                        linked_well_id="W-300")
    rec3 = db.get_procedure(pid)
    ok(rec3.name == "Renamed Proc", "update_procedure name")
    ok(rec3.category_id == cat2, "update_procedure category")
    ok(rec3.linked_well_id == "W-300", "update_procedure keeps links")
    db.close()
    import shutil
    shutil.rmtree(tmp, ignore_errors=True)


def test_reporting():
    print("\n[8] STATISTICAL REPORTING + EXCEL EXPORT")
    import reporting
    p = reporting.procedures_report()
    ok(p["procedures"] > 100, f"procedures > 100 (got {p['procedures']})")
    ok(p["steps"] > 4000, f"steps > 4000 (got {p['steps']})")
    ok(p["hold_points"] >= 0, "hold points counted")
    pr = reporting.problems_report()
    ok(pr["problems"] >= 20, f"problems >= 20 (got {pr['problems']})")
    c = reporting.cbs_report()
    ok(c["items"] >= 300, f"cbs items >= 300 (got {c['items']})")
    ca = reporting.catalog_report()
    ok(ca["docs"] >= 700, f"catalog docs >= 700 (got {ca['docs']})")
    ok(len(ca.get("operation", {})) >= 10, "catalog by operation")
    md = reporting.report_markdown("all")
    ok("STATISTICAL REPORT" in md, "report heading")
    ok("PROCEDURES DATABASE" in md and "KNOWLEDGE LIBRARY" in md,
       "sections present")
    # excel export
    import tempfile
    tmp = tempfile.mkdtemp(prefix="drl_report_")
    xls = os.path.join(tmp, "report.xlsx")
    n = reporting.export_report_excel(xls)
    ok(n >= 7, f"excel sheets >= 7 (got {n})")
    from openpyxl import load_workbook
    wb = load_workbook(xls)
    ok("Library" in wb.sheetnames, "Library sheet present")
    ok(wb["Library"].max_row >= 700, f"Library rows >= 700 "
       f"(got {wb['Library'].max_row})")
    g = reporting.catalog_governance()
    ok(g["docs"] >= 700, "governance counts docs")
    import shutil
    shutil.rmtree(tmp, ignore_errors=True)


def test_neutralize_words():
    print("\n[5] NEUTRALIZE — ordinary English preserved")
    from wizard_engine import neutralize_text
    s = "Total Depth (MD) is 10000 ft; total estimated cost 12M."
    out = neutralize_text(s, "the Operator", "the Service Company")
    ok("Total Depth" in out and "total estimated" in out,
       f"'total' preserved (got: {out!r})")
    s2 = "Per IADC guidelines and API RP 7G."
    out2 = neutralize_text(s2, "the Operator", "the Service Company")
    ok("IADC" in out2 and "API RP 7G" in out2, "'IADC' citation preserved")
    s3 = "Schlumberger and Halliburton must not appear."
    out3 = neutralize_text(s3, "the Operator", "the Service Company")
    ok("Schlumberger" not in out3 and "Halliburton" not in out3,
       "service companies removed")


if __name__ == "__main__":
    test_encrypted_backup()
    test_project_revisions()
    test_secrets()
    test_register()
    test_neutralize_words()
    test_structured_steps()
    test_procedure_linking()
    test_reporting()
    print("\n" + "=" * 60)
    print(f"RESULT: {_PASS} passed, {_FAIL} failed")
    sys.exit(1 if _FAIL else 0)
