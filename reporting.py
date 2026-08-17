# ============================================================================
# STATISTICAL REPORTING & KNOWLEDGE GOVERNANCE
# File: reporting.py
# Audit items:
#   - P2: Central Knowledge Governance (effective-date for documents)
#   - Roadmap: statistical reporting from the databases + Excel export
#
# Produces Markdown + Excel reports over the local databases:
#   procedures (by category / lifecycle status)
#   drilling problems (by category / severity)
#   CBS items (by category, totals)
#   document catalog (754 docs, 5-dimension classification)
#   NPT & lessons learned (operations engine)
#   knowledge governance (effective-date coverage)
# ============================================================================

import json
import os
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

APP_DIR = Path.home() / ".drilling_program"


def _connect(fname: str) -> Optional[sqlite3.Connection]:
    p = APP_DIR / fname
    if not p.exists():
        return None
    con = sqlite3.connect(str(p))
    con.row_factory = sqlite3.Row
    return con


# ---------------------------------------------------------------------------
# Knowledge governance — effective date
# ---------------------------------------------------------------------------

def ensure_effective_date() -> bool:
    """Add the effective_date column to the catalog (idempotent)."""
    con = _connect("catalog.db")
    if con is None:
        return False
    try:
        cols = [r[1] for r in con.execute("PRAGMA table_info(docs)")]
        if "effective_date" not in cols:
            con.execute(
                "ALTER TABLE docs ADD COLUMN effective_date TEXT "
                "DEFAULT '1970-01-01'")
            con.commit()
        return True
    except Exception:
        return False
    finally:
        con.close()


def catalog_governance() -> Dict:
    """Effective-date coverage + per-category statistics of the catalog."""
    ensure_effective_date()
    con = _connect("catalog.db")
    if con is None:
        return {"docs": 0, "dated": 0, "undated": 0, "categories": {}}
    total = con.execute("SELECT COUNT(*) c FROM docs").fetchone()["c"]
    try:
        dated = con.execute(
            "SELECT COUNT(*) c FROM docs WHERE effective_date != "
            "'1970-01-01' AND effective_date != ''").fetchone()["c"]
    except Exception:
        dated = 0
    cats = {}
    for r in con.execute(
            "SELECT category, COUNT(*) c FROM docs GROUP BY category "
            "ORDER BY c DESC"):
        cats[r["category"]] = r["c"]
    con.close()
    return {"docs": total, "dated": dated, "undated": total - dated,
            "categories": cats}


# ---------------------------------------------------------------------------
# Report builders
# ---------------------------------------------------------------------------

def procedures_report() -> Dict:
    con = _connect("procedures.db")
    if con is None:
        return {"procedures": 0}
    out = {"procedures": con.execute(
        "SELECT COUNT(*) c FROM procedures WHERE is_active=1"
    ).fetchone()["c"]}
    out["by_category"] = {}
    for r in con.execute(
            "SELECT COALESCE(c.name,'Uncategorized') cat, COUNT(*) c "
            "FROM procedures p LEFT JOIN categories c "
            "ON p.category_id=c.id WHERE p.is_active=1 "
            "GROUP BY cat ORDER BY c DESC"):
        out["by_category"][r["cat"]] = r["c"]
    out["by_status"] = {}
    for r in con.execute(
            "SELECT COALESCE(status,'Draft') st, COUNT(*) c FROM procedures "
            "WHERE is_active=1 GROUP BY st ORDER BY c DESC"):
        out["by_status"][r["st"]] = r["c"]
    out["steps"] = con.execute(
        "SELECT COUNT(*) c FROM procedure_steps").fetchone()["c"]
    out["checklist"] = con.execute(
        "SELECT COUNT(*) c FROM checklist_items").fetchone()["c"]
    out["hold_points"] = con.execute(
        "SELECT COUNT(*) c FROM procedure_steps WHERE hold_point=1"
    ).fetchone()["c"]
    out["witness_points"] = con.execute(
        "SELECT COUNT(*) c FROM procedure_steps WHERE witness_point=1"
    ).fetchone()["c"]
    out["linked_to_well"] = con.execute(
        "SELECT COUNT(*) c FROM procedures WHERE linked_well_id != ''"
    ).fetchone()["c"]
    con.close()
    return out


def problems_report() -> Dict:
    con = _connect("problems.db")
    if con is None:
        return {"problems": 0}
    out = {"problems": con.execute(
        "SELECT COUNT(*) c FROM problems").fetchone()["c"]}
    out["by_category"] = {}
    out["by_severity"] = {}
    for r in con.execute(
            "SELECT category, COUNT(*) c FROM problems GROUP BY category"):
        out["by_category"][r["category"]] = r["c"]
    for r in con.execute(
            "SELECT severity, COUNT(*) c FROM problems GROUP BY severity"):
        out["by_severity"][r["severity"]] = r["c"]
    con.close()
    return out


def cbs_report() -> Dict:
    con = _connect("cbs.db")
    if con is None:
        return {"items": 0}
    out = {"items": con.execute(
        "SELECT COUNT(*) c FROM cbs_items").fetchone()["c"],
        "priced": con.execute(
            "SELECT COUNT(*) c FROM cbs_items WHERE unit_price>0"
        ).fetchone()["c"]}
    out["total_value"] = round(con.execute(
        "SELECT COALESCE(SUM(unit_price),0) s FROM cbs_items"
    ).fetchone()["s"], 0)
    out["by_category"] = {}
    for r in con.execute(
            "SELECT category, COUNT(*) c, "
            "COALESCE(SUM(unit_price),0) s FROM cbs_items "
            "GROUP BY category ORDER BY s DESC"):
        out["by_category"][r["category"]] = {"count": r["c"],
                                             "value": round(r["s"], 0)}
    con.close()
    return out


def catalog_report() -> Dict:
    ensure_effective_date()
    con = _connect("catalog.db")
    if con is None:
        return {"docs": 0}
    out = {"docs": con.execute(
        "SELECT COUNT(*) c FROM docs").fetchone()["c"]}
    for dim in ("category", "well_type", "environment", "operation"):
        out[dim] = {}
        for r in con.execute(
                f"SELECT {dim} k, COUNT(*) c FROM docs WHERE {dim} != '' "
                f"GROUP BY {dim} ORDER BY c DESC"):
            out[dim][r["k"]] = r["c"]
    con.close()
    return out


def operations_report() -> Dict:
    con = _connect("operations.db")
    if con is None:
        return {"lessons": 0, "npt": 0}
    out = {"lessons": con.execute(
        "SELECT COUNT(*) c FROM lessons").fetchone()["c"],
        "npt": con.execute(
        "SELECT COUNT(*) c FROM npt_events").fetchone()["c"],
        "daily_reports": con.execute(
        "SELECT COUNT(*) c FROM daily_reports").fetchone()["c"]}
    try:
        out["npt_cost"] = round(con.execute(
            "SELECT COALESCE(SUM(direct_cost+indirect_cost),0) s "
            "FROM npt_events").fetchone()["s"], 0)
        out["npt_hours"] = round(con.execute(
            "SELECT COALESCE(SUM(npt_hours),0) s FROM npt_events"
        ).fetchone()["s"], 0)
    except Exception:
        pass
    out["npt_by_cause"] = {}
    try:
        for r in con.execute(
                "SELECT cause, COUNT(*) c FROM npt_events GROUP BY cause "
                "ORDER BY c DESC"):
            out["npt_by_cause"][r["cause"]] = r["c"]
    except Exception:
        pass
    con.close()
    return out


# ---------------------------------------------------------------------------
# Markdown rendering
# ---------------------------------------------------------------------------

def report_markdown(report_type: str = "all") -> str:
    """Assemble a Word-ready statistical report."""
    L = [f"# ENGINEERING & OPERATIONS STATISTICAL REPORT",
         "",
         f"Generated: {datetime.now().strftime('%d-%B-%Y %H:%M')}",
         ""]

    def _table(title, data: Dict, value_fmt=lambda v: str(v)):
        if not data:
            return
        L.append(f"## {title}")
        L.append("")
        L.append("| Category | Count |")
        L.append("|---|---|")
        for k, v in sorted(data.items(), key=lambda kv: -(
                kv[1] if isinstance(kv[1], (int, float)) else 0)):
            L.append(f"| {k} | {value_fmt(v)} |")
        L.append("")

    if report_type in ("all", "procedures"):
        p = procedures_report()
        L.append("## PROCEDURES DATABASE")
        L.append("")
        L.append(f"**{p['procedures']} active procedures** — "
                 f"{p['steps']} steps, {p['checklist']} checklist items, "
                 f"{p['hold_points']} hold points, "
                 f"{p['witness_points']} witness points, "
                 f"{p['linked_to_well']} linked to a well.")
        L.append("")
        _table("By category", p["by_category"])
        _table("By lifecycle status", p["by_status"])

    if report_type in ("all", "problems"):
        pr = problems_report()
        L.append("## DRILLING PROBLEMS KNOWLEDGE BASE")
        L.append("")
        L.append(f"**{pr['problems']} problems**")
        L.append("")
        _table("By category", pr["by_category"])
        _table("By severity", pr["by_severity"])

    if report_type in ("all", "cbs"):
        c = cbs_report()
        L.append("## COST BREAKDOWN STRUCTURE")
        L.append("")
        L.append(f"**{c['items']} items** ({c['priced']} priced) — "
                 f"catalog value {c['total_value']:,.0f} (unit prices).")
        L.append("")
        _table("By category (value)", {k: v["value"] for k, v in
               c["by_category"].items()},
               lambda v: f"{v:,.0f}")

    if report_type in ("all", "catalog"):
        ca = catalog_report()
        L.append("## KNOWLEDGE LIBRARY (DOCUMENT CATALOG)")
        L.append("")
        L.append(f"**{ca['docs']} documents**")
        L.append("")
        _table("By document category", ca.get("category", {}))
        _table("By well type", ca.get("well_type", {}))
        _table("By environment", ca.get("environment", {}))
        _table("By operation", ca.get("operation", {}))

    if report_type in ("all", "operations"):
        o = operations_report()
        L.append("## OPERATIONS REGISTER")
        L.append("")
        L.append(f"**{o['lessons']} lessons learned**, "
                 f"**{o['npt']} NPT events** "
                 f"({o.get('npt_hours', 0):,.0f} hr / "
                 f"{o.get('npt_cost', 0):,.0f} cost), "
                 f"**{o['daily_reports']} daily reports**.")
        L.append("")
        _table("NPT by cause", o.get("npt_by_cause", {}))

    if report_type in ("all", "governance"):
        g = catalog_governance()
        L.append("## KNOWLEDGE GOVERNANCE — EFFECTIVE DATE")
        L.append("")
        L.append(f"**{g['docs']} documents**, {g['dated']} with effective "
                 f"date, {g['undated']} legacy (effective 1970-01-01).")
        L.append("")
    return "\n".join(L)


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def export_report_excel(path: str, report_type: str = "all") -> int:
    """Export the report to a multi-sheet Excel workbook. Returns sheets."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    wb.remove(wb.active)
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="0F3460")

    def sheet(name, rows, headers):
        ws = wb.create_sheet(name[:31])
        for j, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=j, value=h)
            c.font = head_font
            c.fill = head_fill
        for i, r in enumerate(rows, 2):
            for j, v in enumerate(r, 1):
                ws.cell(row=i, column=j, value=v)
        ws.freeze_panes = "A2"
        for j in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + j)].width = 28

    made = 0
    if report_type in ("all", "procedures"):
        p = procedures_report()
        rows = [[k, v] for k, v in p["by_category"].items()]
        sheet("Procedures-Category", rows, ["Category", "Count"])
        rows = [[k, v] for k, v in p["by_status"].items()]
        sheet("Procedures-Status", rows, ["Status", "Count"])
        made += 2
    if report_type in ("all", "problems"):
        pr = problems_report()
        rows = [[k, v] for k, v in pr["by_category"].items()]
        sheet("Problems-Category", rows, ["Category", "Count"])
        rows = [[k, v] for k, v in pr["by_severity"].items()]
        sheet("Problems-Severity", rows, ["Severity", "Count"])
        made += 2
    if report_type in ("all", "cbs"):
        c = cbs_report()
        rows = [[k, v["count"], v["value"]] for k, v in
                c["by_category"].items()]
        sheet("CBS-Category", rows, ["Category", "Items", "Value (USD)"])
        made += 1
    if report_type in ("all", "catalog"):
        ensure_effective_date()
        con = _connect("catalog.db")
        if con:
            rows = [[r["num"], r["title"], r["category"], r["well_type"],
                     r["environment"], r["operation"], r["holes"],
                     r["effective_date"] if "effective_date" in
                     r.keys() else ""]
                    for r in con.execute("SELECT * FROM docs "
                                         "ORDER BY num")]
            con.close()
            sheet("Library", rows, ["Num", "Title", "Category", "Well type",
                                    "Environment", "Operation", "Holes",
                                    "Effective"])
            made += 1
    if report_type in ("all", "operations"):
        o = operations_report()
        rows = [[k, v] for k, v in o.get("npt_by_cause", {}).items()]
        sheet("NPT-Cause", rows, ["Cause", "Events"])
        made += 1

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return made


# ---------------------------------------------------------------------------
# Self-test
# ---------------------------------------------------------------------------

def _selftest():
    ensure_effective_date()
    p = procedures_report()
    assert p["procedures"] > 100, p
    assert p["steps"] > 4000, p
    pr = problems_report()
    assert pr["problems"] >= 20, pr
    c = cbs_report()
    assert c["items"] >= 300, c
    ca = catalog_report()
    assert ca["docs"] >= 700, ca
    md = report_markdown("all")
    assert "STATISTICAL REPORT" in md
    assert "PROCEDURES DATABASE" in md and "KNOWLEDGE LIBRARY" in md
    import tempfile
    tmp = tempfile.mkdtemp(prefix="drl_report_")
    xls = os.path.join(tmp, "report.xlsx")
    n = export_report_excel(xls)
    assert n >= 7, n
    from openpyxl import load_workbook
    wb = load_workbook(xls)
    assert "Library" in wb.sheetnames
    assert wb["Library"].max_row >= 700
    print(f"  ✔ reporting selftest: {p['procedures']} procedures, "
          f"{ca['docs']} docs, {n} sheets")
    return p


if __name__ == "__main__":
    import os as _os
    import sys as _sys
    _sys.path.insert(0, _os.path.dirname(_os.path.abspath(__file__)))
    _selftest()
    print("reporting OK")
